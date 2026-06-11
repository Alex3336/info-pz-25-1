import tkinter as tkr
from tkinter import ttk, messagebox
import urllib.request
import openpyxl
from openpyxl.styles import Alignment
import ctypes
import os

font_url = "https://github.com/google/fonts/raw/main/ufl/ubuntu/Ubuntu-Regular.ttf"
font_filename = "Ubuntu-Regular.ttf"

root = tkr.Tk()
root.withdraw()

if not os.path.exists(font_filename):
    if messagebox.askyesno(
        "Завантаження шрифту",
        "Шрифт 'Ubuntu' не знайдено. Бажаєте завантажити його для кращого вигляду програми?",
    ):
        try:
            print("Завантаження шрифту Ubuntu...")
            urllib.request.urlretrieve(font_url, font_filename)
            print("Завантаження завершено.")
        except Exception as e:
            print(f"Помилка завантаження шрифту: {e}")
    else:
        print(
            "Користувач відмовився від завантаження. Використовується системний шрифт."
        )

if os.name == "nt" and os.path.exists(font_filename):
    ctypes.windll.gdi32.AddFontResourceExW(font_filename, 0x10, 0)
    myappid = "mycompany.myproduct.subproduct.version"
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

root.deiconify()

font_family = "Ubuntu" if os.path.exists(font_filename) else "Arial"


def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = int(screen_width / 2 - width / 2)
    y = int(screen_height / 2 - height / 2)
    window.geometry(f"{width}x{height}+{x}+{y}")


BG_COLOR = "#1f262b"
ACCENT_COLOR = "#4a90e2"
BTN_COLOR = "#4C4281"
TEXT_COLOR = "#f4f4f4"
FONT_MAIN = (font_family, 10)
FONT_BOLD = (font_family, 10, "bold")
FILE_NAME = "hair_services.xlsx"

services = {
    "Олександр Коваль": {
        "чоловіча": {
            "Стрижка": {"duration": 45, "price": 350},
            "Борода": {"duration": 25, "price": 200},
        }
    },
    "Марія Іваненко": {
        "жіноча": {
            "Фарбування": {"duration": 120, "price": 1200},
            "Стрижка": {"duration": 60, "price": 500},
        }
    },
    "Ірина Петренко": {
        "чоловіча": {"Стрижка": {"duration": 50, "price": 300}},
        "жіноча": {"Укладка": {"duration": 40, "price": 450}},
    },
}

root.title("Реєстрація послуги")
root.iconbitmap(default="img/logo_3.ico")
center_window(root, 390, 310)
root.resizable(False, False)
root.configure(bg=BG_COLOR)


style = ttk.Style()
style.theme_use("clam")
style.configure(
    "TCombobox",
    fieldbackground=ACCENT_COLOR,
    background=ACCENT_COLOR,
    foreground=TEXT_COLOR,
    arrowcolor=TEXT_COLOR,
    bordercolor=ACCENT_COLOR,
    lightcolor=ACCENT_COLOR,
    darkcolor=ACCENT_COLOR,
    focuscolor=ACCENT_COLOR,
    borderwidth=0,
    relief=tkr.FLAT,
    padding=0,
)
style.map(
    "TCombobox",
    foreground=[("readonly", TEXT_COLOR)],
    fieldbackground=[("readonly", ACCENT_COLOR)],
    selectforeground=[("readonly", TEXT_COLOR)],
    selectbackground=[("readonly", ACCENT_COLOR)],
)


def refresh(*args):

    selected_sex = sex_var.get()
    selected_service = service_var.get()
    selected_master = master_var.get()

    sexes = set()
    services_list = set()
    masters = set()

    for master, data in services.items():

        for sex, service_data in data.items():

            if selected_sex and sex != selected_sex:
                continue

            for service, info in service_data.items():

                if selected_service and service != selected_service:
                    continue

                if selected_master and master != selected_master:
                    continue

                sexes.add(sex)
                services_list.add(service)
                masters.add(master)

    sex_cb.configure(values=sorted(sexes))
    service_cb.configure(values=sorted(services_list))
    master_cb.configure(values=sorted(masters))

    if selected_sex not in sexes:
        sex_var.set("")

    if selected_service not in services_list:
        service_var.set("")

    if selected_master not in masters:
        master_var.set("")

    if (
        selected_master
        and selected_service
        and selected_sex
        and selected_master in services
        and selected_sex in services[selected_master]
        and selected_service in services[selected_master][selected_sex]
    ):

        info = services[selected_master][selected_sex][selected_service]

        duration_var.set(f'{info["duration"]} хв')
        price_var.set(f'{info["price"]} грн')

    else:

        duration_var.set("")
        price_var.set("")


def update_info(*args):

    sex = sex_var.get()
    service = service_var.get()
    master = master_var.get()

    if not master:
        return

    info = services[master][sex][service]

    duration_var.set(f'{info["duration"]} хв')

    price_var.set(f'{info["price"]} грн')


def save_data():

    sex = sex_var.get()
    service = service_var.get()
    master = master_var.get()
    duration = duration_entry.get()
    price = price_entry.get()

    if not all([sex, service, duration, master, price]):
        messagebox.showwarning("Помилка", "Заповніть усі поля")
        return

    record = {
        "Стать": sex,
        "Послуга": service,
        "Майстер": master,
        "Тривалість": duration,
        "Вартість": price,
    }

    try:
        if os.path.exists(FILE_NAME):
            workbook = openpyxl.load_workbook(FILE_NAME)
            service_sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            service_sheet = workbook.active
            service_sheet.append(list(record.keys()))

            service_sheet.append(list(record.values()))

            for col in service_sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    cell.alignment = Alignment(horizontal="left")
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                service_sheet.column_dimensions[column_letter].width = max_length + 2

            for row in service_sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="left")

            workbook.save(FILE_NAME)

            messagebox.showinfo("Успіх", f"Дані збережено до {FILE_NAME}!)")
    except Exception as e:
        messagebox.showerror("Помилка", f"Не вдалося зберегти файл: {e}")

    clear_fields()


def clear_fields():

    sex_cb.set("")

    master_cb["values"] = []
    master_cb.set("")

    service_cb["values"] = []
    service_cb.set("")

    duration_var.set("")
    price_var.set("")


def open_xlsx():

    if not os.path.exists(FILE_NAME):
        messagebox.showerror("Помилка", "Файл ще не створений")
        return

    os.startfile(FILE_NAME)


global_frame = tkr.LabelFrame(
    root,
    text="Вхідні дані:",
    bg=BG_COLOR,
    fg=ACCENT_COLOR,
    font=FONT_BOLD,
    relief=tkr.RIDGE,
    bd=2,
)
global_frame.place(y=20, x=20, width=330, height=250)

labels = ["Стать:", "Назва послуги:", "ПІБ майстра:", "Тривалість:", "Вартість:"]

for i, text in enumerate(labels):
    tkr.Label(global_frame, text=text, bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_MAIN).grid(
        row=i + 1, column=0, sticky="w", padx=20, pady=5
    )

sex_var = tkr.StringVar()
service_var = tkr.StringVar()
master_var = tkr.StringVar()

sex_cb = ttk.Combobox(
    global_frame, textvariable=sex_var, values=["чоловіча", "жіноча"], state="readonly"
)
service_cb = ttk.Combobox(global_frame, textvariable=service_var, state="readonly")
master_cb = ttk.Combobox(global_frame, textvariable=master_var, state="readonly")

sex_cb.grid(row=1, column=1, padx=20, sticky="we")
service_cb.grid(row=2, column=1, padx=20, sticky="ew")
master_cb.grid(row=3, column=1, padx=20, sticky="ew")

duration_var = tkr.StringVar()
price_var = tkr.StringVar()

duration_entry = tkr.Entry(global_frame, textvariable=duration_var, state="readonly")
duration_entry.configure(readonlybackground=ACCENT_COLOR, fg=TEXT_COLOR)
duration_entry.grid(row=4, column=1, padx=20, sticky="ew")

price_entry = tkr.Entry(global_frame, textvariable=price_var, state="readonly")
price_entry.configure(readonlybackground=ACCENT_COLOR, fg=TEXT_COLOR)
price_entry.grid(row=5, column=1, padx=20, sticky="ew")


save_btn = tkr.Button(
    global_frame,
    text="Зберегти",
    bg=BTN_COLOR,
    fg=TEXT_COLOR,
    command=save_data,
    relief=tkr.FLAT,
    cursor="hand2",
)
save_btn.place(y=180, x=22, width=85)

reset_button = tkr.Button(
    global_frame,
    text="Очистити",
    bg=BTN_COLOR,
    fg=TEXT_COLOR,
    command=clear_fields,
    relief=tkr.FLAT,
    cursor="hand2",
)
reset_button.place(y=180, x=117, width=85)

open_button = tkr.Button(
    global_frame,
    text="Відкрити xlsx",
    bg=BTN_COLOR,
    fg=TEXT_COLOR,
    command=open_xlsx,
    relief=tkr.FLAT,
    cursor="hand2",
)
open_button.place(y=180, x=212, width=85)


sex_var.trace_add("write", refresh)
service_var.trace_add("write", refresh)
master_var.trace_add("write", refresh)


author = tkr.Label(
    root,
    text="\u00a9 Шарандак О.В., ПЗ-25-1",
    bg=BG_COLOR,
    fg=ACCENT_COLOR,
    font=(font_family, 8, "italic"),
)
author.place(x=20, y=280)

refresh()

root.mainloop()
